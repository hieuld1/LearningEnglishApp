import pygame
import sys
import openpyxl
import os
import requests

from PIL import Image
from io import BytesIO
from configure_box import ConfigBox
from button import Button
from googletrans import Translator

class CreateLesson:
    def __init__(self, ai_game):

        self.screen = ai_game.screen
        self.screen_rect = self.screen.get_rect()
        self.font = pygame.font.Font(None, 36)

        self.ConfigBoxLesson    = ConfigBox(self, 250, 60, 140, 32, 'Lesson')
        self.ConfigBoxWord      = ConfigBox(self, 250, 160, 140, 32, 'Word')
        self.ButtonCreate       = Button(self, 250, 100, "Create")
        self.ButtonInsert       = Button(self, 250, 200, "Insert")
        self.ButtonFinish       = Button(self, 250, 260, "Finish")

        self.excel_file_name            = 'test.xlsx'
        self.excel_file_name_no_tail    = 'test'
                
        # Flag handle learn configure done
        self.learn_configure_done   = False

        self.find_word_fail         = False
        self.find_image_fail        = False

    def create_lesson_set_config(self, val):
        self.learn_configure_done = val

    def create_lesson_get_config(self):
        return self.learn_configure_done   

    def create_lesson_draw(self):
        # print("draw lesson windows")
        self.ButtonFinish.draw_button()
        self.ButtonInsert.draw_button()
        self.ButtonCreate.draw_button()

        self.ConfigBoxLesson.config_box_display()
        self.ConfigBoxWord.config_box_display()

        if self.find_word_fail:
            text = "Can't translate word"
            text_render = self.font.render(text, True, (30, 30, 30))
            self.screen.blit(text_render, (400, 160))

        if self.find_image_fail:
            text = "Can't find image"
            text_render = self.font.render(text, True, (30, 30, 30))
            self.screen.blit(text_render, (400, 200))

    def create_lesson_handle_event(self, event):
        if event.type == pygame.MOUSEBUTTONDOWN:
            mouse_pos = pygame.mouse.get_pos()
            self._check_finish_button(mouse_pos)
            self._check_insert_button(mouse_pos)
            self._check_create_button(mouse_pos)

        self.ConfigBoxLesson.handle_event(event)
        self.ConfigBoxWord.handle_event(event)

    def _check_finish_button(self, mouse_pos):
        """Start a new game when the player clicks Play."""
        button_clicked = self.ButtonFinish.rect.collidepoint(mouse_pos)
        
        if button_clicked:
            print("Button Finish")

            # Flag handle learn configure done
            self.learn_configure_done    = False

    def _check_insert_button(self, mouse_pos):
        """Start a new game when the player clicks Play."""
        button_clicked = self.ButtonInsert.rect.collidepoint(mouse_pos)
        
        if button_clicked:
            print("Button insert") 
            word1 = str(self.ConfigBoxWord.text)
            word2 = self.__translate_to_vietnamese(word1)
            
            if word2 != None:
                self._insert_excel_file(word1, word2)  
                self.__get_image(word1)

            self.ConfigBoxWord.text = ''  

    def _check_create_button(self, mouse_pos):
        """Start a new file xlsx to create lesson"""
        button_clicked = self.ButtonCreate.rect.collidepoint(mouse_pos)
        
        if button_clicked:
            print("Button create")
            file_name = str(self.ConfigBoxLesson.text)
            self.open_or_create_excel_file(file_name)
            # self._create_excel_file(file_name + '.xlsx')

    # def _create_excel_file(self, file_name):
    #     # Create a new workbook
    #     self.workbook = openpyxl.Workbook()

    #     # Save file name
    #     self.workbook.save(file_name)

    #     self.excel_file_name = file_name

    def _insert_excel_file(self, data1, data2):
        # Load excel file
        self.workbook = openpyxl.load_workbook(self.excel_file_name)

        # Select te active sheet
        sheet = self.workbook.active

        print("max row:", sheet.max_row)
        rows_idx = sheet.max_row + 1

        # Insert data
        sheet.cell(row = rows_idx, column = 1, value = data1) 
        sheet.cell(row = rows_idx, column = 2, value = data2) 

        # Save change
        self.workbook.save(self.excel_file_name)

    def __translate_to_vietnamese(self, word):
        translator = Translator()
        translation = translator.translate(word, src='en', dest='vi')
        print("Word2: ", translation.text)
        if translation.text != word:
            self.find_word_fail = False
            return translation.text
        else:
            print("Can't translate word: ", word);
            self.find_word_fail = True
            return None
    
    def open_or_create_excel_file(self, file_name):

        folder_path = "./"
        file_path = os.path.join(folder_path, f"{file_name}.xlsx")

        if os.path.exists(file_path):
            # If the file exists, open it
            workbook = openpyxl.load_workbook(file_path)
            print(f"Opening existing file: {file_name}.xlsx")
        else:
            # If the file doesn't exist, create a new one
            workbook = openpyxl.Workbook()
            workbook.save(file_path)
            print(f"Creating a new file: {file_name}.xlsx")

        self.excel_file_name = file_name + '.xlsx'
        self.excel_file_name_no_tail = file_name

        return workbook
    
    def __get_image(self, word):
        # Replace 'YOUR_UNSPLASH_ACCESS_KEY' with your actual Unsplash Access Key
        access_key = 'hOk85mGaPnka79ncCAuCi6agDY6UWChTMuVwMNmx6k0'
        
        # Make a request to the Unsplash API to get a random image based on the input word
        response = requests.get(f'https://api.unsplash.com/photos/random?query={word}&client_id={access_key}')
        
        if response.status_code == 200:
            # Extract the image URL from the API response
            image_url = response.json()['urls']['regular']
            
            # Download the image
            image_data = requests.get(image_url).content
            
            # Open the image using PIL
            img = Image.open(BytesIO(image_data))
            
            # Resize the image to 240 x 160 pixels
            img = img.resize((240, 160))
            
            # Save the image to the 'images' folder
            img.save(f'images/{self.excel_file_name_no_tail}_{word}.jpg')
            self.find_image_fail = False
        else:
            print("Can't find image E: %", response.status_code)
            self.find_image_fail = True
            

    def show_error_window(self, error_message):
        pygame.init()
        error_screen = pygame.display.set_mode((300, 200))
        pygame.display.set_caption("Error Window")
        
        font = pygame.font.Font(None, 36)
        text_surface = font.render(error_message, True, (255, 0, 0))
        text_rect = text_surface.get_rect(center=(150, 100))
        
        error_screen.fill((255, 255, 255))
        error_screen.blit(text_surface, text_rect)
        pygame.display.flip()
        
        while True:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    return  # Return to main window if error window is closed