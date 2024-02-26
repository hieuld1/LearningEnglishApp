# -*- coding: utf-8 -*-
import openpyxl

class ReadExcelFile:
    """Read excel file to get data."""

    def __init__(self):
        """Initialize read excel file."""
        wb = openpyxl.load_workbook('English3000.xlsx')
        sheet = wb.active

        self.list_key = []
        self.list_value = []
        self.word_dictionary = {}

        for i in range(1, sheet.max_row):
            # print(i, sheet.cell(row = i, column = 1).value, sheet.cell(row = i, column = 3).value)
            
            key_read = sheet.cell(row = i, column = 1).value
            string_read = self.__get_first_word(str(key_read))
            self.list_key = self.list_key + [string_read]

            val_read = sheet.cell(row = i, column = 3).value
            self.list_value = self.list_value + [val_read]

        # print(self.list_key)
        # print(self.list_value)

        for key, value in zip(self.list_key, self.list_value):
            self.word_dictionary[key] = value

        # print(list_key[0])
            
    def getKey(self, index):
        return str(self.list_key[index])   

    def getValue(self, index):
        return str(self.list_value[index])    

    def __get_first_word(self, input_string):
        return input_string.split() [0] 

if __name__ == '__main__':
    # Make a game instance, and run the game.
    ai = ReadExcelFile()