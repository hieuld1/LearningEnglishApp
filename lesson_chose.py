import pygame
import sys
import openpyxl
import os

from tkinter import Tk, filedialog
from button import Button

from settings import Settings
from LearnWindows import LearnWin
from inputBox import InputBox

LEARN_MODE_LESSON_CARD      = 2
LEARN_MODE_LESSON_WORD      = 3
NUM_LEARN_WORDS_DONE        = 1000

class LessonChoose:

    def __init__(self, ai_game):

        self.screen = ai_game.screen
        self.screen_rect = self.screen.get_rect()
        self.font = pygame.font.Font(None, 36)

        # Make the Play button.
        self.ButtonOpenFile     = Button(self, 200, 160, "Lesson")
        self.yes_button         = Button(self, 400, 260, "Yes")
        self.no_button          = Button(self, 0, 260, "No")
        self.flip_button        = Button(self, 200, 0, "Flip")
        self.settings           = Settings()
        self.LearnWin           = LearnWin(self)
        self.InputBox           = InputBox(self, 140, 60, 140, 32)

        # Create a learning list
        self.learn_word_key     = []
        self.learn_word_value   = []
        self.learn_word_stat    = []
        self.display_word       = ''
        self.display_word_idx   = 0
        self.display_word_key   = 1     # Flash card 1; Fill word 0
        self.learning_number_word       = 0

        self.learning_mode              = LEARN_MODE_LESSON_CARD

        # Load the word image and get its rect.
        self.learn_word_image           = pygame.image.load('images/test_an.jpg')
        self.learn_word_image_rect      = self.learn_word_image.get_rect()

        # Start each new word  at the bottom center of the screen.
        self.learn_word_image_rect.midleft = self.screen_rect.midleft

        self.learn_file_name            = 'test'

        # Flag handle learn choose done
        self.lesson_choose_done         = False

        # Flag handle learn configure done
        self.learn_configure_done       = False

    def choose_lesson_set_config(self, val):
        self.learn_configure_done = val

    def choose_lesson_get_config(self):
        return self.learn_configure_done  

    def __choose_lesson_get_file_name(self, file_path):
        # Split the file path into directory and file name
        directory, file_name = os.path.split(file_path)

        # Split the file name and extension
        file_name, file_extension = os.path.splitext(file_name)

        return file_name

    def choose_lesson_draw_picture(self, word):  
        self.learn_word_image   = pygame.image.load(f'images/{self.learn_file_name}_{word}.jpg')
        self.screen.blit(self.learn_word_image, self.learn_word_image_rect)

    def choose_lesson_draw(self):
        # print("draw lesson windows")
        if self.lesson_choose_done:
            self.LearnWin.show_word(str(self.display_word))
            self.yes_button.draw_button()
            self.no_button.draw_button()
            self.flip_button.draw_button()

            if(self.learning_mode == LEARN_MODE_LESSON_CARD):
                self.choose_lesson_draw_picture(self.display_word)

            if(self.learning_mode == LEARN_MODE_LESSON_WORD):
                # print("Lesson word")
                if self.display_word_idx < self.learning_number_word:
                    correct_word = self.learn_word_key[self.display_word_idx]
                    # print("Word: ", correct_word)
                    self.InputBox.input_box_display(correct_word)
        else:
            self.ButtonOpenFile.draw_button()

    def choose_lesson_handle_event(self, event):
        if event.type == pygame.MOUSEBUTTONDOWN:
            mouse_pos = pygame.mouse.get_pos()
            if self.lesson_choose_done:
                self._check_flip_button(mouse_pos)
                self._check_yes_button(mouse_pos)
                self._check_no_button(mouse_pos)

            else:
                self._check_open_file_button(mouse_pos)

        if self.lesson_choose_done:
            if(self.learning_mode == LEARN_MODE_LESSON_WORD):
                self.InputBox.handle_event(event)

    def _check_open_file_button(self, mouse_pos):
        """Start a new game when the player clicks Play."""
        button_clicked = self.ButtonOpenFile.rect.collidepoint(mouse_pos)
        
        if button_clicked:
            print("Button open file")
            self.choose_lesson_open_file_dialog()

    def _check_flip_button(self, mouse_pos):
        """Start a new game when the player clicks Play."""
        button_clicked = self.flip_button.rect.collidepoint(mouse_pos)
        
        if button_clicked:
            if self.display_word_key:
                self.display_word_key   = 0  
                self.display_word       = self.learn_word_value[self.display_word_idx]
            else:
                self.display_word_key   = 1  
                self.display_word       = self.learn_word_key[self.display_word_idx]
            
            # print(self.display_word) 
    
    def _check_yes_button(self, mouse_pos):
        """Start a new game when the player clicks Play."""
        button_clicked = self.yes_button.rect.collidepoint(mouse_pos)
        
        if button_clicked:
            # print("BTYes: ", self.learning_number_word, self.display_word_idx)
            if self.display_word_idx < self.learning_number_word:
                self.learn_word_stat[self.display_word_idx] = 1

            self.display_word_idx   = self.__find_next_word()

            if self.display_word_idx < self.learning_number_word:
                if(self.learning_mode == LEARN_MODE_LESSON_CARD):
                    self.display_word_key   = 1 
                    self.display_word       = self.learn_word_key[self.display_word_idx]
                else:
                    self.display_word_key   = 0 
                    self.display_word       = self.learn_word_value[self.display_word_idx]
                
            # print(self.display_word)  


    def _check_no_button(self, mouse_pos):
        """Start a new game when the player clicks Play."""
        button_clicked = self.no_button.rect.collidepoint(mouse_pos)
        
        if button_clicked:
            self.display_word_idx   = self.__find_next_word()

            if self.display_word_idx < self.learning_number_word:
                if(self.learning_mode == LEARN_MODE_LESSON_CARD):
                    self.display_word_key   = 1 
                    self.display_word       = self.learn_word_key[self.display_word_idx]
                else:
                    self.display_word_key   = 0 
                    self.display_word       = self.learn_word_value[self.display_word_idx]
                
            # print(self.display_word)
                    
    def __find_next_word(self):
        
        num_loop = 0
        index_out = self.display_word_idx 

        while True:
            index_out   = index_out + 1
            
            if(index_out >= self.learning_number_word):
                index_out = 0
                
            # print(index_out, self.learning_number_word)
            if self.learn_word_stat[index_out] == 0:
                return index_out

            num_loop = num_loop + 1

            if(num_loop >= self.learning_number_word):
                print("All words have been learned!")
                # self.display_word       = 'finish work!'
                # Flag handle learn configure/choose done
                self.lesson_choose_done     = False
                self.learn_configure_done   = False
                
                return NUM_LEARN_WORDS_DONE

    def choose_lesson_open_file_dialog(self):
        
        root = Tk()
        root.withdraw()  # Hide the main window

        file_path = filedialog.askopenfilename(title="Select a File", filetypes = [("Excel file", "*.xlsx")])

        self.learn_file_name            = self.__choose_lesson_get_file_name(file_path)
        print(self.learn_file_name )

        if file_path:
            # print("Selected File:", file_path)
            self.choose_lesson_read_excel(file_path)
            # print("Read done")
            self.lesson_choose_done = True

            # print("Number word: ", self.learning_number_word)
            if self.learning_number_word > 0:
                self.display_word_idx       = 0
                if(self.learning_mode == LEARN_MODE_LESSON_CARD):
                    self.display_word_key   = 1 
                    self.display_word       = self.learn_word_key[0]
                else:
                    self.display_word_key   = 0 
                    self.display_word       = self.learn_word_value[0]

            # print("display word: ", self.display_word)

    def choose_lesson_read_excel(self, file_path):
        # print("read excel file")
        try:
            workbook    = openpyxl.load_workbook(file_path)
            sheet       = workbook.active

            # print(sheet.max_row)

            for i in range(2, sheet.max_row + 1):
                print(i, sheet.cell(row = i, column = 1).value, sheet.cell(row = i, column = 2).value)
                key_read = sheet.cell(row = i, column = 1).value
                # string_read = self.__get_first_word(str(key_read))
                self.learn_word_key = self.learn_word_key + [key_read]

                val_read = sheet.cell(row = i, column = 2).value
                self.learn_word_value = self.learn_word_value + [val_read]

                self.learn_word_stat    = self.learn_word_stat + [0]

                # image
                # self.learn_word_image   = self.learn_word_image + [pygame.image.load(f'images/{self.learn_file_name}_{key_read}.jpg')]

            workbook.close()

            self.learning_number_word = len(self.learn_word_value)
            print("Number word: ", self.learning_number_word)

        except Exception as e:
            print(f"Error: {str(e)}")
    
    def choose_lesson_set_learn_mode(self, learn_mode):
        self.learning_mode = learn_mode
