# import pygame
# import sys

# class InputBox:
#     def __init__(self, x, y, width, height, font):
#         self.rect = pygame.Rect(x, y, width, height)
#         self.color_inactive = pygame.Color('lightskyblue3')
#         self.color_active = pygame.Color('dodgerblue2')
#         self.color = self.color_inactive
#         self.text = ''
#         self.font = font
#         self.active = False

#     def handle_event(self, event):
#         if event.type == pygame.MOUSEBUTTONDOWN:
#             if self.rect.collidepoint(event.pos):
#                 self.active = not self.active
#             else:
#                 self.active = False
#             self.color = self.color_active if self.active else self.color_inactive

#         if event.type == pygame.KEYDOWN:
#             if self.active:
#                 if event.key == pygame.K_RETURN:
#                     return True
#                 elif event.key == pygame.K_BACKSPACE:
#                     self.text = self.text[:-1]
#                 else:
#                     self.text += event.unicode
#         return False

#     def update(self):
#         width = max(200, self.font.size(self.text)[0] + 10)
#         self.rect.w = width

#     def draw(self, screen):
#         txt_surface = self.font.render(self.text, True, self.color)
#         width = max(200, txt_surface.get_width()+10)
#         self.rect.w = width
#         screen.blit(txt_surface, (self.rect.x+5, self.rect.y+5))
#         pygame.draw.rect(screen, self.color, self.rect, 2)

# # Initialize Pygame
# pygame.init()

# # Set up display
# width, height = 400, 200
# screen = pygame.display.set_mode((width, height))
# pygame.display.set_caption("String Comparison")

# # Set up fonts
# font = pygame.font.Font(None, 36)

# # Set up colors
# white = (255, 255, 255)
# black = (0, 0, 0)

# # Set up strings
# correct_string = "Python"
# user_input = ""

# # Create InputBox instance
# input_box = InputBox(50, 50, 140, 32, font)

# # Main game loop
# while True:
#     for event in pygame.event.get():
#         if event.type == pygame.QUIT:
#             pygame.quit()
#             sys.exit()

#         input_box.handle_event(event)

#     # Clear the screen
#     screen.fill(white)

#     # Update and draw InputBox
#     input_box.update()
#     input_box.draw(screen)

#     # Compare input with correct string
#     if input_box.text == correct_string:
#         result_text = "Correct!"
#     else:
#         result_text = "Wrong!"

#     # Render and draw result text
#     result_text_render = font.render(result_text, True, black)
#     screen.blit(result_text_render, (50, 100))

#     # Update the display
#     pygame.display.flip()

#     # Set the frames per second
#     pygame.time.Clock().tick(60)


# import sys
# import pygame
 
# pygame.init()
# screen = pygame.display.set_mode((400, 400))
# clock = pygame.time.Clock()
 
# text_font = pygame.font.SysFont('arial', 15)
# text = text_font.render("àéèù φχψω kết giao, liên kết, kết hợp, cho cộng tác", True, (0,0,0))

# while True:
#     for event in pygame.event.get():
#         if event.type == pygame.QUIT:
#             pygame.quit()
#             sys.exit()

#     screen.fill((255, 255, 255))
#     screen.blit(text, (40, 40))

#     pygame.display.flip()
#     clock.tick(60)

# import pygame
# import sys

# # Initialize Pygame
# pygame.init()

# # Set up display
# width, height = 400, 300
# screen = pygame.display.set_mode((width, height))
# pygame.display.set_caption("Text Wrapping in Pygame")

# # Set up font
# font_size = 20
# font = pygame.font.Font(None, font_size)

# # Custom text wrapping function
# def draw_text(surface, text, font, color, rect):
#     words = text.split()
#     space_width, _ = font.size(' ')
#     x, y = rect.topleft
#     line_spacing = font.get_linesize()

#     for word in words:
#         word_width, word_height = font.size(word)

#         if x + word_width > rect.right:
#             x = rect.left
#             y += line_spacing

#         surface.blit(font.render(word, True, color), (x, y))
#         x += word_width + space_width

# # Main loop
# while True:
#     for event in pygame.event.get():
#         if event.type == pygame.QUIT:
#             pygame.quit()
#             sys.exit()

#     # Clear the screen
#     screen.fill((255, 255, 255))

#     # Your text to be wrapped
#     text_to_wrap = "This is a long text that needs to be wrapped in the display area of the Pygame window."

#     # Define the rect area for text
#     text_rect = pygame.Rect(50, 50, width - 100, height - 100)

#     # Call the custom function to wrap and draw text
#     draw_text(screen, text_to_wrap, font, (0, 0, 0), text_rect)

#     # Update the display
#     pygame.display.flip()

# # Quit Pygame
# pygame.quit()

# def __filter_list_letter(input_list, input_letter):
#     output_list = []
#     for string in input_list:
#         if string.lower().startswith(input_letter.lower()):
#             output_list = output_list + [string]

#     print(output_list)

#     return output_list

# input = ["abc", "bcd", "cde"]
# input_letter = "a"

# # __filter_list_letter(input, input_letter)
# print(len(input))


# from googletrans import Translator

# def translate_to_vietnamese(word):
#     translator = Translator()
#     translation = translator.translate(word, src='en', dest='vi')
#     return translation.text

# # Example usage:
# english_word = "fantastic"
# vietnamese_translation = translate_to_vietnamese(english_word)
# print(f"{english_word} in Vietnamese: {vietnamese_translation}")

# import subprocess
# import platform

# def enable_ethernet():
#     if platform.system().lower() == 'windows':
#         subprocess.run(['netsh', 'interface', 'set', 'interface', 'Ethernet', 'admin=enable'], capture_output=True)
#         print("Ethernet connection enabled.")
#     else:
#         print("This script is intended for Windows only.")

# def disable_ethernet():
#     if platform.system().lower() == 'windows':
#         subprocess.run(['netsh', 'interface', 'set', 'interface', 'Ethernet', 'admin=disable'], capture_output=True)
#         print("Ethernet connection disabled.")
#     else:
#         print("This script is intended for Windows only.")

# # Example usage
# enable_ethernet()  # Uncomment this line to enable Ethernet
# # disable_ethernet()  # Uncomment this line to disable Ethernet

# ====================================================================
import pygame
import sys

class PygameListBox:
    def __init__(self, width, height, options):
        pygame.init()

        self.width = width
        self.height = height
        self.options = options

        self.screen = pygame.display.set_mode((width, height))
        pygame.display.set_caption('Pygame List Box')

        self.font = pygame.font.Font(None, 36)
        self.selected_index = None

        self.run()

    def run(self):
        while True:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    sys.exit()
                elif event.type == pygame.MOUSEBUTTONDOWN:
                    if event.button == 1:  # Left mouse button
                        x, y = event.pos
                        self.handle_click(x, y)

            self.draw()

    def handle_click(self, x, y):
        item_height = self.height // len(self.options)

        clicked_index = y // item_height
        if 0 <= clicked_index < len(self.options):
            self.selected_index = clicked_index
            print(f"Selected Option: {self.options[self.selected_index]}")
            # Perform action based on the selected option

    def draw(self):
        self.screen.fill((255, 255, 255))

        item_height = self.height // len(self.options)
        for i, option in enumerate(self.options):
            rect = pygame.Rect(0, i * item_height, self.width, item_height)
            color = (200, 200, 200) if i != self.selected_index else (150, 150, 150)
            pygame.draw.rect(self.screen, color, rect)
            text_surface = self.font.render(option, True, (0, 0, 0))
            text_rect = text_surface.get_rect(center=(self.width // 2, i * item_height + item_height // 2))
            self.screen.blit(text_surface, text_rect)

        pygame.display.flip()

# if __name__ == "__main__":
#     options = ["Flash Cards", "Fill Words", "Create Lessons"]
#     PygameListBox(400, 300, options)

# ============================================================================

import pygame
import sys

class ComboBox:
    def __init__(self, options, x, y, width, height, font_size=20):
        self.options = options
        self.selected_option = None
        self.rect = pygame.Rect(x, y, width, height)
        self.font = pygame.font.Font(None, font_size)
        self.is_open = False

    def draw(self, screen):
        pygame.draw.rect(screen, (200, 200, 200), self.rect)
        pygame.draw.polygon(screen, (0, 0, 0), [(self.rect.x + self.rect.width - 20, self.rect.centery - 5),
                                                (self.rect.x + self.rect.width - 10, self.rect.centery - 5),
                                                (self.rect.x + self.rect.width - 15, self.rect.centery + 5)])

        selected_text = self.font.render(self.selected_option if self.selected_option else "Select Option", True, (0, 0, 0))
        screen.blit(selected_text, (self.rect.x + 10, self.rect.y + self.rect.height // 2 - selected_text.get_height() // 2))

        if self.is_open:
            for i, option in enumerate(self.options):
                option_rect = pygame.Rect(self.rect.x, self.rect.y + self.rect.height * (i + 1), self.rect.width, self.rect.height)
                pygame.draw.rect(screen, (200, 200, 200), option_rect)
                text = self.font.render(option, True, (0, 0, 0))
                screen.blit(text, (option_rect.x + 10, option_rect.y + option_rect.height // 2 - text.get_height() // 2))

    def handle_event(self, event):
        if event.type == pygame.MOUSEBUTTONDOWN:
            if event.button == 1:
                if self.rect.collidepoint(event.pos):
                    self.is_open = not self.is_open
                elif self.is_open:
                    for i, option in enumerate(self.options):
                        option_rect = pygame.Rect(self.rect.x, self.rect.y + self.rect.height * (i + 1),
                                                 self.rect.width, self.rect.height)
                        if option_rect.collidepoint(event.pos):
                            self.selected_option = option
                            self.is_open = False

# # Pygame initialization
# pygame.init()

# # Screen setup
# screen_width, screen_height = 400, 300
# screen = pygame.display.set_mode((screen_width, screen_height))
# pygame.display.set_caption("Pygame ComboBox")

# # ComboBox setup
# options_list = ["Option 1", "Option 2", "Option 3"]
# combo_box = ComboBox(options_list, 50, 50, 200, 30)

# # Main loop
# while True:
#     for event in pygame.event.get():
#         if event.type == pygame.QUIT:
#             pygame.quit()
#             sys.exit()
#         combo_box.handle_event(event)

#     screen.fill((255, 255, 255))
#     combo_box.draw(screen)

#     pygame.display.flip()
                            
# text = 'abcd'                           
# print((str(text) + '.xlsx'))
                            
# import os
# import openpyxl

# def open_or_create_excel_file(folder_path, file_name):
#     file_path = os.path.join(folder_path, f"{file_name}.xlsx")

#     if os.path.exists(file_path):
#         # If the file exists, open it
#         workbook = openpyxl.load_workbook(file_path)
#         print(f"Opening existing file: {file_name}.xlsx")
#     else:
#         # If the file doesn't exist, create a new one
#         workbook = openpyxl.Workbook()
#         workbook.save(file_path)
#         print(f"Creating a new file: {file_name}.xlsx")

#     return workbook

# # Example usage:
# folder_path = "./"
# input_file_name = input("Enter the file name (without extension): ")

# # Call the function
# workbook_instance = open_or_create_excel_file(folder_path, input_file_name)


import tkinter as tk
from tkinter import filedialog
import openpyxl

def open_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        read_excel(file_path)

def read_excel(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=1, values_only=True):
            print(row)

        workbook.close()
    except Exception as e:
        print(f"Error: {str(e)}")

# # Create the main window
# root = tk.Tk()
# root.title("XLSX Reader")

# # Create and configure the Open File button
# open_button = tk.Button(root, text="Open File", command=open_file)
# open_button.pack(pady=20)

# # Start the Tkinter event loop
# root.mainloop()

import pygame
from tkinter import Tk, filedialog

class Game:
    def __init__(self):
        pygame.init()
        self.screen = pygame.display.set_mode((400, 200))
        pygame.display.set_caption("Pygame with File Dialog")

    def open_file_dialog(self):
        root = Tk()
        root.withdraw()  # Hide the main window

        file_path = filedialog.askopenfilename(title="Select a File")
        if file_path:
            print("Selected File:", file_path)

    def run(self):
        while True:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    quit()
                elif event.type == pygame.MOUSEBUTTONDOWN:
                    self.open_file_dialog()

            pygame.display.flip()

# if __name__ == "__main__":
#     game = Game()
#     game.run()

# import requests
# from PIL import Image

# def get_image_url(api_key, query):
#     url = f'https://api.unsplash.com/photos/random'
#     headers = {'Authorization': f'Client-ID {api_key}'}
#     params = {'query': query}
#     response = requests.get(url, headers=headers, params=params)
#     data = response.json()
#     return data['urls']['regular']

# def display_image(url):
#     image = Image.open(requests.get(url, stream=True).raw)
#     image.show()

# def main():
#     api_key = 'hOk85mGaPnka79ncCAuCi6agDY6UWChTMuVwMNmx6k0'
#     word = input("Enter a word: ")
#     image_url = get_image_url(api_key, word)
#     display_image(image_url)

# if __name__ == "__main__":
#     main()
            
# import requests
# from PIL import Image
# from io import BytesIO

# def get_image(word):
#     # Replace 'YOUR_UNSPLASH_ACCESS_KEY' with your actual Unsplash Access Key
#     access_key = 'hOk85mGaPnka79ncCAuCi6agDY6UWChTMuVwMNmx6k0'
    
#     # Make a request to the Unsplash API to get a random image based on the input word
#     response = requests.get(f'https://api.unsplash.com/photos/random?query={word}&client_id={access_key}')
    
#     # Extract the image URL from the API response
#     image_url = response.json()['urls']['regular']
    
#     # Download the image
#     image_data = requests.get(image_url).content
    
#     # Open the image using PIL
#     img = Image.open(BytesIO(image_data))
    
#     # Resize the image to 240 x 160 pixels
#     img = img.resize((240, 160))
    
#     # Save the image to the 'images' folder
#     img.save(f'images/{word}_image.jpg')

# # Example usage
# word_input = input("Enter a word: ")
# get_image(word_input)
            




import sqlite3
from datetime import datetime

def create_table(conn):
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS program_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            start_time TEXT,
            exit_time TEXT,
            learn_lesson TEXT
        )
    ''')
    conn.commit()

def save_program_execution(conn):
    start_time1 = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Simulate program execution
    input("Press Enter when the program is about to exit...")

    learn_lesson1 = input("Input lesson to learn...")

    exit_time1 = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO program_records (start_time, exit_time, learn_lesson) VALUES (?, ?, ?)
    ''', (start_time1, exit_time1, learn_lesson1))
    conn.commit()

def read_program_records(conn):
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM program_records')
    records = cursor.fetchall()

    for record in records:
        print(f"Start Time: {record[1]}, Exit Time: {record[2]}, learn_lesson: {record[3]}")

if __name__ == "__main__":
    # Connect to SQLite database (change the database name if using a different database)
    connection = sqlite3.connect('program_records.db')

    # Create the table if it doesn't exist
    create_table(connection)

    # Save program execution record
    save_program_execution(connection)

    # Read and display all program records
    read_program_records(connection)

    # Close the database connection
    connection.close()





